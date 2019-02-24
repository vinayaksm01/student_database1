from tkinter import *
from PIL import ImageTk, Image
import os
import xlrd
import openpyxl

root = Tk()
img = ImageTk.PhotoImage(Image.open("C:/Users/chethan.R/PycharmProjects/student_database/std_data/person-984236_960_720.jpg"))
panel = Label(root, image = img)
panel.grid(row=1,column=1)
def get_data():
    def farecal():
        e1 = l1.get()
        excel_file = 'C:/Users/chethan.R/PycharmProjects/student_database/excel_data/id.xlsx'

        wb = xlrd.open_workbook(excel_file)
        sheet = wb.sheet_by_index(0)

        sheet.cell_value(1,1)
        for i in range(sheet.nrows):
            if e1 == sheet.cell_value(i,0):
                # x=sheet.row_values(0)
                # d=sheet.row_values(i)
                a = sheet.row_values(i,0,1)

                b = sheet.row_values(i,1,2)
                c = sheet.row_values(i,2,3)

                d = sheet.row_values(i,3,4)
                # e = sheet.row_values(i,4,5)

                # e3.insert(END,x)
                # e4.insert(END,d)
                e3.insert(END, a)
                e4.insert(END, b)
                e5.insert(END, c)
                e6.insert(END, d)
                # e7.insert(END,e)

    root = Tk()
    root.title("")
    root.geometry("900x900")
    lable3 = Label(root, text="WELCOME TO STUDENT DATA", fg="white", bg="blue", width=100, height=2)
    e3 = Text(root, width=50, height=2)
    e4: Text = Text(root, width=50, height=2)
    e5: Text = Text(root, width=50, height=2)
    e6: Text = Text(root, width=50, height=2)
    # e7: Text = Text(root, width=50, height=2)

    lable1 = Label(root, text="ENTER YOUR USN")
    lable4 = Label(root, text="STUDENT USN")
    lable5 = Label(root, text="STUDENT NAME")
    lable6 = Label(root, text="STUDENT BRANCH")
    lable7 = Label(root, text="STUDENT ADDRESS")
    b1 = Button(root, width=30, text="Get Student Details", command=farecal)

    l1 = Entry(root)
    lable3.grid(row=0, column=1)
    lable1.grid(row=1, column=0)
    l1.grid(row=1, column=1)
    b1.grid(row=3, column=1)

    e3.grid(row=6, column=1)
    lable4.grid(row=6, column=0)
    lable5.grid(row=8, column=0)
    lable6.grid(row=10, column=0)
    lable7.grid(row=12, column=0)
    e4.grid(row=8, column=1)
    e5.grid(row=10, column=1)
    e6.grid(row=12, column=1)
    # e7.grid(row=14,column=1)
    root.mainloop()




def add_data():
    def save():
        l3 = e3.get()
        l4 = e4.get()
        l5 = e5.get()
        l6 = e6.get()
        wb = openpyxl.load_workbook(filename='C:/Users/chethan.R/PycharmProjects/student_database/excel_data/id.xlsx')
        ws = wb.worksheets[0]
        c = [l3,l4,l5,l6]
        row_count = ws.max_row + 1
        for i in range(len(c)):
            # if ws.cell(1,1).value == xlrd.empty_cell.value:
            # ab=ws.cell(row=i+1,column=1)

            column_count = ws.max_column
            ws.cell(row_count, column=i + 1).value = c[i]
            wb.save("C:/Users/chethan.R/PycharmProjects/student_database/excel_data/id.xlsx")

    root = Tk()
    root.title("")
    root.geometry("900x900")
    lable3 = Label(root, text="WELCOME TO STUDENT DATA", fg="white", bg="blue", width=100, height=2)
    e3 = Entry(root,width=50)
    e4 = Entry(root,width=50)
    e5 = Entry(root,width=50)
    e6 = Entry(root,width=50)
    # e3 = Text(root, width=50, height=2)
    # e4 = Text(root, width=50, height=2)
    # e5 = Text(root, width=50, height=2)
    # e6 = Text(root, width=50, height=2)
    # # e7: Text = Text(root, width=50, height=2)
    lable4 = Label(root, text="STUDENT USN")
    lable5 = Label(root, text="STUDENT NAME")
    lable6 = Label(root, text="STUDENT BRANCH")
    lable7 = Label(root, text="STUDENT ADDRESS")
    b1 = Button(root, width=30, text="Add Details", command=save)


    lable3.grid(row=0, column=1)

    b1.grid(row=14, column=1)

    e3.grid(row=6, column=1)

    lable4.grid(row=6, column=0)
    lable5.grid(row=8, column=0)
    lable6.grid(row=10, column=0)
    lable7.grid(row=12, column=0)
    e4.grid(row=8, column=1)
    e5.grid(row=10, column=1)
    e6.grid(row=12, column=1)
    # e7.grid(row=14,column=1)
    root.mainloop()
# B1 = tkinter.Button(top, text ="Add Student details", command = add_data)
# B2 = tkinter.Button(top, text ="Get Student Details", command = std_data)
#
# top.geometry("900x900")


# img = ImageTk.PhotoImage(Image.open("girl-students-l.jpg"))
# top = Label(top, image = img)
# top.pack(side = "bottom", fill = "both", expand = "yes")
# top.mainloop()

#
# B1.grid(row=3,column=0)
# B2.grid(row=4,column=0)
# top.mainloop()
b1 = Button(root, width=60, text="Add Student Details", command=add_data)
b2 = Button(root, width=60, text="Get Student Details", command=get_data)
b1.grid(row=3,column=1)
b2.grid(row=4,column=1)

root.mainloop()